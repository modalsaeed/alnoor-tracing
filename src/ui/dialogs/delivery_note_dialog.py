"""
Delivery Note Dialog - Generate delivery notes from verified coupons.
"""

from datetime import datetime
from pathlib import Path
from typing import List, Optional
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QPushButton, QLabel, QComboBox, QDateEdit, QSpinBox,
    QGroupBox, QMessageBox, QFileDialog, QTableWidget,
    QTableWidgetItem, QHeaderView
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QFont

from src.database.db_manager import DatabaseManager
from src.database.models import PatientCoupon, MedicalCentre, DistributionLocation, Product, PurchaseOrder, DeliveryNote
from src.utils import Colors, StyleSheets
from src.utils.model_helpers import get_attr, get_id, get_name, get_nested_attr


class DeliveryNoteDialog(QDialog):
    """Dialog for generating delivery notes from coupons."""
    
    def __init__(self, db_manager: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.filtered_coupons = []
        self.selected_product = None
        self.selected_centre = None
        self.selected_location = None
        
        self.setWindowTitle("Generate Delivery Note")
        self.setModal(True)
        self.setMinimumWidth(900)
        self.setMinimumHeight(700)
        
        self.setup_ui()
        self.load_filter_data()
    
    def setup_ui(self):
        """Set up the user interface."""
        layout = QVBoxLayout(self)
        
        # Title
        title = QLabel("📄 Generate Delivery Note")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        layout.addWidget(title)
        
        # Info message
        info = QLabel(
            "Select filters to find unverified coupons, then generate a delivery note.\n"
            "The delivery note will be saved as an Excel file."
        )
        info.setStyleSheet(
            "background-color: #e3f2fd; padding: 10px; border-radius: 4px; "
            "color: #1565c0; border-left: 4px solid #2196f3;"
        )
        info.setWordWrap(True)
        layout.addWidget(info)
        
        layout.addSpacing(15)
        
        # Filters section
        filters_group = QGroupBox("Filters (Select coupons to include)")
        filters_layout = QFormLayout()
        
        # Health Centre filter
        self.centre_combo = QComboBox()
        self.centre_combo.currentIndexChanged.connect(self.apply_filters)
        filters_layout.addRow("MOH Health Centre: *", self.centre_combo)
        
        # Distribution Location filter
        self.location_combo = QComboBox()
        self.location_combo.currentIndexChanged.connect(self.apply_filters)
        filters_layout.addRow("Distribution Location:", self.location_combo)
        
        # Purchase Order filter
        self.po_combo = QComboBox()
        filters_layout.addRow("PO Reference: *", self.po_combo)
        
        # Date range
        date_layout = QHBoxLayout()
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_from.dateChanged.connect(self.apply_filters)
        date_layout.addWidget(self.date_from)
        
        date_layout.addWidget(QLabel("to"))
        
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.dateChanged.connect(self.apply_filters)
        date_layout.addWidget(self.date_to)
        
        filters_layout.addRow("Date Range:", date_layout)
        
        filters_group.setLayout(filters_layout)
        layout.addWidget(filters_group)
        
        # Preview section
        preview_group = QGroupBox("Coupon Preview (Unverified only)")
        preview_layout = QVBoxLayout()
        
        # Summary label
        self.summary_label = QLabel("No coupons match the selected filters")
        self.summary_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        preview_layout.addWidget(self.summary_label)
        
        # Table
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(6)
        self.preview_table.setHorizontalHeaderLabels([
            "Coupon Ref", "Patient Name", "CPR", "Product", "Quantity", "Date Received"
        ])
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.preview_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setMaximumHeight(250)
        preview_layout.addWidget(self.preview_table)
        
        preview_group.setLayout(preview_layout)
        layout.addWidget(preview_group)
        
        # Delivery details section
        details_group = QGroupBox("Delivery Note Details")
        details_layout = QFormLayout()
        
        # Pieces per carton
        self.pieces_per_carton_input = QSpinBox()
        self.pieces_per_carton_input.setRange(1, 10000)
        self.pieces_per_carton_input.setValue(1)
        self.pieces_per_carton_input.setSuffix(" pieces")
        details_layout.addRow("Pieces per Carton: *", self.pieces_per_carton_input)
        
        # Calculated values display
        self.total_pieces_label = QLabel("0")
        self.total_pieces_label.setStyleSheet("font-weight: bold; color: #27ae60;")
        details_layout.addRow("Total Pieces:", self.total_pieces_label)
        
        self.total_cartons_label = QLabel("0")
        self.total_cartons_label.setStyleSheet("font-weight: bold; color: #27ae60;")
        details_layout.addRow("Total Cartons:", self.total_cartons_label)
        
        details_group.setLayout(details_layout)
        layout.addWidget(details_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_btn)
        
        self.generate_btn = QPushButton("📄 Generate Delivery Note")
        self.generate_btn.clicked.connect(self.generate_delivery_note)
        self.generate_btn.setStyleSheet(StyleSheets.button_primary(Colors.SUCCESS))
        self.generate_btn.setEnabled(False)
        button_layout.addWidget(self.generate_btn)
        
        layout.addLayout(button_layout)
    
    def load_filter_data(self):
        """Load data for filters."""
        try:
            # Load health centres
            centres = sorted(self.db_manager.get_all(MedicalCentre), key=lambda x: get_name(x))
            self.centre_combo.addItem("-- Select Health Centre --", None)
            for centre in centres:
                self.centre_combo.addItem(get_name(centre), get_id(centre))
            
            # Load distribution locations
            locations = sorted(self.db_manager.get_all(DistributionLocation), key=lambda x: get_name(x))
            self.location_combo.addItem("-- Select Distribution Location --", None)
            for location in locations:
                self.location_combo.addItem(get_name(location), get_id(location))
            
            # Load purchase orders
            pos = sorted(self.db_manager.get_all(PurchaseOrder), key=lambda x: get_attr(x, 'po_reference', ''))
            self.po_combo.addItem("-- Select PO Reference --", None)
            for po in pos:
                display_text = f"{get_attr(po, 'po_reference', 'N/A')}"
                product_name = get_nested_attr(po, 'product.name')
                if product_name:
                    display_text += f" ({product_name})"
                self.po_combo.addItem(display_text, get_id(po))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load filter data: {str(e)}")
    
    def apply_filters(self):
        """Apply filters and update preview."""
        centre_id = self.centre_combo.currentData()
        location_id = self.location_combo.currentData()
        date_from = self.date_from.date().toPyDate()
        date_to = self.date_to.date().toPyDate()
        
        # Clear previous results
        self.filtered_coupons = []
        self.selected_product = None
        self.selected_centre = None
        self.selected_location = None
        
        # Centre must be selected, location is optional
        if not centre_id:
            self.update_preview()
            return
        
        try:
            # Get all coupons and filter
            all_coupons = self.db_manager.get_all(PatientCoupon)
            date_from_dt = datetime.combine(date_from, datetime.min.time())
            date_to_dt = datetime.combine(date_to, datetime.max.time())
            
            coupons = []
            import datetime as dt
            for c in all_coupons:
                if get_attr(c, 'verified', True) != False:
                    continue
                if get_attr(c, 'medical_centre_id') != centre_id:
                    continue
                date_val = get_attr(c, 'date_received')
                coupon_dt = None
                if isinstance(date_val, dt.datetime):
                    coupon_dt = date_val
                elif isinstance(date_val, str):
                    try:
                        coupon_dt = dt.datetime.fromisoformat(date_val)
                    except Exception:
                        continue
                if not coupon_dt:
                    continue
                if not (date_from_dt <= coupon_dt <= date_to_dt):
                    continue
                if location_id and get_attr(c, 'distribution_location_id') != location_id:
                    continue
                coupons.append(c)
            
            # Group by product
            product_groups = {}
            for coupon in coupons:
                product_id = get_attr(coupon, 'product_id')
                if product_id:
                    if product_id not in product_groups:
                        product_groups[product_id] = []
                    product_groups[product_id].append(coupon)
            
            # For simplicity, we'll handle one product at a time
            # If multiple products, warn the user
            if len(product_groups) > 1:
                QMessageBox.warning(
                    self,
                    "Multiple Products",
                    f"Found coupons for {len(product_groups)} different products.\n"
                    "Delivery notes are generated per product.\n"
                    "Please generate separate delivery notes for each product."
                )
                self.filtered_coupons = []
            elif len(product_groups) == 1:
                product_id = list(product_groups.keys())[0]
                self.filtered_coupons = product_groups[product_id]
                all_products = self.db_manager.get_all(Product)
                all_centres = self.db_manager.get_all(MedicalCentre)
                all_locations = self.db_manager.get_all(DistributionLocation)
                self.selected_product = next((p for p in all_products if get_id(p) == product_id), None)
                self.selected_centre = next((c for c in all_centres if get_id(c) == centre_id), None)
                self.selected_location = next((l for l in all_locations if get_id(l) == location_id), None) if location_id else None
            else:
                self.filtered_coupons = []
            
            self.update_preview()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to filter coupons: {str(e)}")
    
    def update_preview(self):
        """Update the preview table and summary."""
        # Update table
        self.preview_table.setRowCount(0)
        
        total_pieces = 0
        for coupon in self.filtered_coupons:
            row = self.preview_table.rowCount()
            self.preview_table.insertRow(row)
            coupon_ref = get_attr(coupon, 'coupon_reference', 'N/A')
            patient_name = get_attr(coupon, 'patient_name', 'N/A')
            cpr = get_attr(coupon, 'cpr', 'N/A')
            # Product name: try to get product object or fallback to product_id
            product = get_attr(coupon, 'product', None)
            if product:
                product_name = get_attr(product, 'name', 'N/A')
            else:
                # Try to resolve from selected_product or fallback to product_id
                product_id = get_attr(coupon, 'product_id', None)
                if self.selected_product and get_id(self.selected_product) == product_id:
                    product_name = get_attr(self.selected_product, 'name', 'N/A')
                else:
                    product_name = str(product_id) if product_id else 'N/A'
            quantity_pieces = get_attr(coupon, 'quantity_pieces', 0)
            date_val = get_attr(coupon, 'date_received', None)
            date_str = 'N/A'
            if date_val:
                import datetime as dt
                if isinstance(date_val, dt.datetime):
                    date_str = date_val.strftime("%d/%m/%Y")
                elif isinstance(date_val, str):
                    try:
                        date_obj = dt.datetime.fromisoformat(date_val)
                        date_str = date_obj.strftime("%d/%m/%Y")
                    except Exception:
                        date_str = date_val
            self.preview_table.setItem(row, 0, QTableWidgetItem(str(coupon_ref)))
            self.preview_table.setItem(row, 1, QTableWidgetItem(str(patient_name)))
            self.preview_table.setItem(row, 2, QTableWidgetItem(str(cpr)))
            self.preview_table.setItem(row, 3, QTableWidgetItem(str(product_name)))
            self.preview_table.setItem(row, 4, QTableWidgetItem(str(quantity_pieces)))
            self.preview_table.setItem(row, 5, QTableWidgetItem(str(date_str)))
            total_pieces += quantity_pieces
        
        # Update summary
        if self.filtered_coupons:
            product_name = get_attr(self.selected_product, 'name', 'N/A') if self.selected_product else 'N/A'
            self.summary_label.setText(
                f"Found {len(self.filtered_coupons)} coupon(s) | "
                f"Product: {product_name} | "
                f"Total: {total_pieces} pieces"
            )
            self.generate_btn.setEnabled(True)
        else:
            self.summary_label.setText("No coupons match the selected filters")
            self.generate_btn.setEnabled(False)
        
        # Update calculated values
        self.total_pieces_label.setText(str(total_pieces))
        pieces_per_carton = self.pieces_per_carton_input.value()
        total_cartons = total_pieces / pieces_per_carton if pieces_per_carton > 0 else 0
        self.total_cartons_label.setText(f"{total_cartons:.2f}")
        
        # Connect pieces per carton change
        self.pieces_per_carton_input.valueChanged.connect(self.update_carton_calculation)
    
    def update_carton_calculation(self):
        """Update carton calculation when pieces per carton changes."""
        total_pieces = sum(get_attr(c, 'quantity_pieces', 0) for c in self.filtered_coupons)
        pieces_per_carton = self.pieces_per_carton_input.value()
        total_cartons = total_pieces / pieces_per_carton if pieces_per_carton > 0 else 0
        self.total_cartons_label.setText(f"{total_cartons:.2f}")
    
    def generate_delivery_note(self):
        """Generate the delivery note Excel file and save a DeliveryNote record."""
        if not self.filtered_coupons:
            QMessageBox.warning(self, "No Data", "No coupons selected for delivery note.")
            return
        if not self.selected_product or not self.selected_centre:
            QMessageBox.warning(self, "Missing Data", "Product or Health Centre information is missing.")
            return
        po_id = self.po_combo.currentData()
        if not po_id:
            QMessageBox.warning(self, "Missing PO", "Please select a PO reference.")
            return
        try:
            all_pos = self.db_manager.get_all(PurchaseOrder)
            selected_po = next((po for po in all_pos if get_id(po) == po_id), None)
            if not selected_po:
                QMessageBox.warning(self, "Error", "Selected PO not found.")
                return
            po_reference = get_attr(selected_po, 'po_reference', 'N/A')
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to get PO reference: {str(e)}")
            return
        centre_name = get_attr(self.selected_centre, 'name', 'Centre').replace(' ', '_') if self.selected_centre else 'Centre'
        default_filename = f"DeliveryNote_{centre_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Delivery Note",
            str(Path.home() / "Documents" / default_filename),
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            import sys
            if getattr(sys, 'frozen', False):
                base_path = Path(sys._MEIPASS)
            else:
                base_path = Path(__file__).parent.parent.parent.parent
            template_path = base_path / "resources" / "templates" / "delivery_note_template.xlsx"
            if template_path.exists():
                wb = load_workbook(template_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
            # Generate delivery note number (now based on DeliveryNote count)
            all_notes = self.db_manager.get_all(DeliveryNote)
            max_number = 0
            for note in all_notes:
                dn = get_attr(note, 'delivery_note_number', None)
                if dn and dn.startswith("DNM-"):
                    try:
                        num = int(dn.split("-")[1])
                        max_number = max(max_number, num)
                    except (IndexError, ValueError):
                        pass
            next_number = max_number + 1
            dn_number = f"DNM-{next_number:05d}"
            total_pieces = sum(get_attr(c, 'quantity_pieces', 0) for c in self.filtered_coupons)
            pieces_per_carton = self.pieces_per_carton_input.value()
            total_cartons = total_pieces / pieces_per_carton
            ws['C2'] = f"Delivery Note: {dn_number}"
            ws['C3'] = f"Ref: {centre_name}"
            ws['C4'] = f"Ref: PO-{po_reference}"
            ws['E2'] = f"Date: {datetime.now().strftime('%d-%b-%Y')}"
            product_ref = get_attr(self.selected_product, 'reference', '') if self.selected_product else ''
            ws['C13'] = product_ref
            product_name = get_attr(self.selected_product, 'name', 'N/A') if self.selected_product else 'N/A'
            ws['C14'] = product_name
            ws['F14'] = total_pieces
            ws['E14'] = pieces_per_carton
            ws['D14'] = total_cartons
            ws['F20'] = total_pieces
            wb.save(file_path)
            # Save DeliveryNote record
            dn_data = {
                'delivery_note_number': dn_number,
                'centre_id': get_id(self.selected_centre),
                'centre_name': get_attr(self.selected_centre, 'name', ''),
                'location_id': get_id(self.selected_location) if self.selected_location else None,
                'location_name': get_attr(self.selected_location, 'name', '') if self.selected_location else '',
                'product_id': get_id(self.selected_product),
                'product_name': get_attr(self.selected_product, 'name', ''),
                'po_id': get_id(selected_po),
                'po_reference': po_reference,
                'total_pieces': total_pieces,
                'total_cartons': total_cartons,
                'date_created': datetime.now(),
                'file_path': file_path
            }
            new_note = self.db_manager.create(DeliveryNote, dn_data)
            # (Removed) No longer updating coupons to reference this delivery note
            QMessageBox.information(
                self,
                "Success",
                f"Delivery note generated successfully!\n\n"
                f"Delivery Note: {dn_number}\n"
                f"File saved to: {file_path}\n\n"
                f"Total Coupons: {len(self.filtered_coupons)}\n"
                f"Total Pieces: {total_pieces}\n"
                f"Total Cartons: {total_cartons:.2f}"
            )
            if self.parent() and hasattr(self.parent(), 'load_recent_delivery_notes'):
                self.parent().load_recent_delivery_notes()
            self.accept()
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to generate delivery note:\n{str(e)}"
            )
    
    def generate_delivery_note_number(self) -> str:
        """Generate auto-incremented delivery note number."""
        try:
            with self.db_manager.get_session() as session:
                # Get the highest delivery note number
                coupons = session.query(PatientCoupon).filter(
                    PatientCoupon.delivery_note_number.isnot(None)
                ).all()
                
                max_number = 0
                for coupon in coupons:
                    dn = coupon.delivery_note_number
                    if dn and dn.startswith("DNM-"):
                        try:
                            num = int(dn.split("-")[1])
                            max_number = max(max_number, num)
                        except (IndexError, ValueError):
                            pass
                
                next_number = max_number + 1
                return f"DNM-{next_number:05d}"
        except Exception as e:
            print(f"Error generating delivery note number: {e}")
            # Fallback to timestamp-based number
            return f"DNM-{int(datetime.now().timestamp())}"
    
    def update_coupon_delivery_notes(self, dn_number: str):
        """Update the delivery note number for all coupons in this batch."""
        try:
            is_api = hasattr(self.db_manager, 'is_api_client') and getattr(self.db_manager, 'is_api_client', False)
            if is_api:
                # API mode: update each coupon via db_manager.update
                for coupon in self.filtered_coupons:
                    update_data = dict(getattr(coupon, '__dict__', {})) if hasattr(coupon, '__dict__') else dict(coupon)
                    update_data['delivery_note_number'] = dn_number
                    self.db_manager.update(PatientCoupon, get_id(coupon), update_data)
            else:
                with self.db_manager.get_session() as session:
                    for coupon in self.filtered_coupons:
                        db_coupon = session.query(PatientCoupon).get(coupon.id)
                        if db_coupon:
                            db_coupon.delivery_note_number = dn_number
                    session.commit()
        except Exception as e:
            print(f"Error updating coupon delivery notes: {e}")
