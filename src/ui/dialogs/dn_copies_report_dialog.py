"""
Delivery Notes Copies Received Report Dialog
"""

from datetime import datetime
from pathlib import Path
from typing import List, Dict
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QPushButton, QLabel, QComboBox, QDateEdit, QSpinBox,
    QGroupBox, QMessageBox, QFileDialog, QCheckBox
)
from PyQt6.QtCore import Qt, QDate

from src.database.db_manager import DatabaseManager
from src.database.models import PatientCoupon, MedicalCentre, PurchaseOrder
from src.utils import Colors, StyleSheets
from src.utils.model_helpers import get_attr, get_id, get_name, get_nested_attr


class DNcopiesReportDialog(QDialog):
    """Dialog for generating Delivery Notes Copies Received report."""
    
    def __init__(self, db_manager: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        
        self.setWindowTitle("Generate DN Copies Report")
        self.setModal(True)
        self.setMinimumWidth(600)
        
        self.setup_ui()
        self.load_po_data()
    
    def setup_ui(self):
        """Set up the user interface."""
        layout = QVBoxLayout(self)
        
        # Title
        title = QLabel("📋 Delivery Notes Copies Received Report")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        layout.addWidget(title)
        
        # Info message
        info = QLabel(
            "Generate a report of delivery notes with verification numbers and GRV references.\n"
            "Select the PO reference, date range, and pieces per carton to calculate totals."
        )
        info.setStyleSheet(
            "background-color: #e3f2fd; padding: 10px; border-radius: 4px; "
            "color: #1565c0; border-left: 4px solid #2196f3;"
        )
        info.setWordWrap(True)
        layout.addWidget(info)
        
        layout.addSpacing(15)
        
        # Parameters section
        params_group = QGroupBox("Report Parameters")
        params_layout = QFormLayout()
        
        # PO Reference
        self.po_combo = QComboBox()
        params_layout.addRow("PO Reference: *", self.po_combo)
        
        # Date range
        date_layout = QHBoxLayout()
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        date_layout.addWidget(self.date_from)
        
        date_layout.addWidget(QLabel("to"))
        
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        date_layout.addWidget(self.date_to)
        
        params_layout.addRow("Date Range:", date_layout)
        
        # Pieces per carton
        self.pieces_per_carton = QSpinBox()
        self.pieces_per_carton.setRange(1, 10000)
        self.pieces_per_carton.setValue(160)
        self.pieces_per_carton.setSuffix(" pieces")
        params_layout.addRow("Pieces per Carton: *", self.pieces_per_carton)
        
        # GRV filter
        self.grv_filter = QCheckBox("Only include delivery notes with GRV")
        params_layout.addRow("GRV Filter:", self.grv_filter)
        
        params_group.setLayout(params_layout)
        layout.addWidget(params_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        generate_btn = QPushButton("📄 Generate Report")
        generate_btn.clicked.connect(self.generate_report)
        generate_btn.setStyleSheet(StyleSheets.button_primary(Colors.SUCCESS))
        button_layout.addWidget(generate_btn)
        
        layout.addLayout(button_layout)
    
    def load_po_data(self):
        """Load purchase orders for selection."""
        try:
            pos = self.db_manager.get_all(PurchaseOrder)
            pos = sorted(pos, key=lambda po: get_attr(po, 'po_reference', ''))
            self.po_combo.addItem("-- Select PO Reference --", None)
            for po in pos:
                display_text = f"{get_attr(po, 'po_reference', '')}"
                product = get_attr(po, 'product', None)
                product_name = get_attr(product, 'name', '') if product else ''
                if product_name:
                    display_text += f" ({product_name})"
                self.po_combo.addItem(display_text, get_id(po))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load PO data: {str(e)}")
    
    def generate_report(self):
        """Generate the DN Copies report."""
        # Validate inputs
        po_id = self.po_combo.currentData()
        if not po_id:
            QMessageBox.warning(self, "Missing PO", "Please select a PO reference.")
            return
        
        date_from = self.date_from.date().toPyDate()
        date_to = self.date_to.date().toPyDate()
        pieces_per_carton = self.pieces_per_carton.value()
        grv_filter = self.grv_filter.isChecked()
        
        try:
            from datetime import datetime as dt
            # Always use db_manager.get_all and Python-side filtering for both API and ORM
            all_pos = self.db_manager.get_all(PurchaseOrder)
            po = next((p for p in all_pos if get_id(p) == po_id), None)
            if not po:
                QMessageBox.warning(self, "Error", "Selected PO not found.")
                return
            po_reference = get_attr(po, 'po_reference', 'N/A')
            all_coupons = self.db_manager.get_all(PatientCoupon)
            date_from_dt = dt.combine(date_from, dt.min.time())
            date_to_dt = dt.combine(date_to, dt.max.time())
            coupons = []
            for c in all_coupons:
                dn_num = get_attr(c, 'delivery_note_number', None)
                if not dn_num:
                    continue
                if not get_attr(c, 'date_received', None):
                    continue
                # Date filter
                date_val = get_attr(c, 'date_received', None)
                coupon_dt = None
                if isinstance(date_val, dt):
                    coupon_dt = date_val
                elif isinstance(date_val, str):
                    try:
                        coupon_dt = dt.fromisoformat(date_val)
                    except Exception:
                        continue
                if not coupon_dt or not (date_from_dt <= coupon_dt <= date_to_dt):
                    continue
                # GRV filter
                if grv_filter:
                    grv_ref = get_attr(c, 'grv_reference', None)
                    if not grv_ref:
                        continue
                coupons.append(c)
            if not coupons:
                QMessageBox.information(
                    self, 
                    "No Data", 
                    "No delivery notes found matching the selected criteria."
                )
                return
            # Group coupons by delivery note number
            dn_groups = {}
            for coupon in coupons:
                dn_num = get_attr(coupon, 'delivery_note_number', None)
                if dn_num not in dn_groups:
                    dn_groups[dn_num] = []
                dn_groups[dn_num].append(coupon)
            # Prepare DN data
            dn_data = []
            # Preload all medical centres for id lookup
            all_centres = self.db_manager.get_all(MedicalCentre)
            def resolve_centre_name(coupon, po):
                # 1. Try coupon.medical_centre.name
                name = get_nested_attr(coupon, 'medical_centre.name', None)
                if name and name != 'Unknown':
                    return name
                # 2. Try coupon.medical_centre_id
                centre_id = get_attr(coupon, 'medical_centre_id', None)
                if centre_id:
                    for centre in all_centres:
                        if get_id(centre) == centre_id:
                            return get_name(centre)
                # 3. Try PO.medical_centre.name
                name = get_nested_attr(po, 'medical_centre.name', None)
                if name and name != 'Unknown':
                    return name
                # 4. Try PO.medical_centre_id
                centre_id = get_attr(po, 'medical_centre_id', None)
                if centre_id:
                    for centre in all_centres:
                        if get_id(centre) == centre_id:
                            return get_name(centre)
                return 'Unknown'
            for dn_number, dn_coupons in dn_groups.items():
                first_coupon = dn_coupons[0]
                centre_name = resolve_centre_name(first_coupon, po)
                # Get date received (use earliest date if multiple)
                date_received = None
                for c in dn_coupons:
                    date_val = get_attr(c, 'date_received', None)
                    coupon_dt = None
                    if isinstance(date_val, dt):
                        coupon_dt = date_val
                    elif isinstance(date_val, str):
                        try:
                            coupon_dt = dt.fromisoformat(date_val)
                        except Exception:
                            continue
                    if not date_received or (coupon_dt and coupon_dt < date_received):
                        date_received = coupon_dt
                verification_ref = get_attr(first_coupon, 'verification_reference', '-')
                grv_ref = get_attr(first_coupon, 'grv_reference', '-')
                total_pieces = sum(get_attr(c, 'quantity_pieces', 0) for c in dn_coupons)
                total_cartons = total_pieces / pieces_per_carton
                dn_data.append({
                    'dn_number': dn_number,
                    'centre_name': centre_name,
                    'date_received': date_received,
                    'verification_ref': verification_ref,
                    'grv_ref': grv_ref,
                    'total_pieces': total_pieces,
                    'total_cartons': total_cartons
                })
            dn_data.sort(key=lambda x: x['date_received'] or dt.min)
            self.create_excel_report(po_reference, dn_data, pieces_per_carton)
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to generate report:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
    
    def create_excel_report(self, po_reference: str, dn_data: List[Dict], pieces_per_carton: int):
        """Create the Excel report file."""
        # Ask user where to save
        default_filename = f"DN_Copies_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save DN Copies Report",
            str(Path.home() / "Documents" / default_filename),
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import load_workbook
            
            # Load template
            template_path = Path(__file__).parent.parent.parent.parent / "resources" / "templates" / "delivery_note_copies_template.xlsx"
            
            if not template_path.exists():
                QMessageBox.critical(
                    self,
                    "Template Missing",
                    f"Template file not found at:\n{template_path}\n\n"
                    "Please place the template file in the correct location."
                )
                return
            
            wb = load_workbook(template_path)
            ws = wb.active
            
            # Fill B11: PO reference as 'PO-{ref}'
            ws['B11'] = f"PO-{po_reference}"
            
            # Fill A12: Current date as 'Date: {date}'
            ws['A12'] = f"Date: {datetime.now().strftime('%d-%b-%Y')}"
            
            # Fill data starting from row 15
            from copy import copy
            
            current_row = 15
            total_cartons = 0
            total_pieces = 0
            
            for idx, dn in enumerate(dn_data, start=1):
                # Copy row 15 formatting to new rows if needed
                if idx > 1:
                    # Insert a new row by copying row 15
                    ws.insert_rows(current_row)
                    # Copy formatting from row 15 to the new row
                    for col in range(1, 9):  # A to H
                        source_cell = ws.cell(15, col)
                        target_cell = ws.cell(current_row, col)
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.number_format = copy(source_cell.number_format)
                            target_cell.protection = copy(source_cell.protection)
                            target_cell.alignment = copy(source_cell.alignment)
                
                ws[f'A{current_row}'] = idx
                ws[f'B{current_row}'] = dn['centre_name']
                ws[f'C{current_row}'] = dn['dn_number']
                ws[f'D{current_row}'] = dn['date_received'].strftime('%d-%b-%Y')
                ws[f'E{current_row}'] = dn['verification_ref']
                ws[f'F{current_row}'] = f"{dn['total_cartons']:.2f}"
                ws[f'G{current_row}'] = dn['total_pieces']
                ws[f'H{current_row}'] = dn['grv_ref']
                
                total_cartons += dn['total_cartons']
                total_pieces += dn['total_pieces']
                current_row += 1
            
            # Add totals row
            ws[f'E{current_row}'] = "TOTAL"
            ws[f'F{current_row}'] = f"{total_cartons:.2f}"
            ws[f'G{current_row}'] = total_pieces
            
            # Save workbook
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "Success",
                f"DN Copies report generated successfully!\n\n"
                f"File saved to: {file_path}\n\n"
                f"Total Delivery Notes: {len(dn_data)}\n"
                f"Total Cartons: {total_cartons:.2f}\n"
                f"Total Pieces: {total_pieces}"
            )
            
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to create Excel file:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
