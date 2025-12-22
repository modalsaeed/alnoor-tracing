"""
Script to create the delivery note Excel template.
Run this once to generate the template file.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path

def create_delivery_note_template():
    """Create the delivery note Excel template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Note"
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Define styles
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=11)
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title - Row 1
    ws['A1'] = 'DELIVERY NOTE'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.merge_cells('A1:F1')
    
    # Delivery Note Number - C2 (will be filled by program)
    ws['C2'] = 'Delivery Note: '
    ws['C2'].font = header_font
    ws['C2'].alignment = left_align
    
    # Reference - C3 (will be filled by program)
    ws['C3'] = 'Ref: '
    ws['C3'].font = normal_font
    ws['C3'].alignment = left_align
    
    # Date - E2 (will be filled by program)
    ws['E2'] = 'Date: '
    ws['E2'].font = normal_font
    ws['E2'].alignment = left_align
    
    # Empty rows for spacing
    ws.row_dimensions[4].height = 5
    ws.row_dimensions[5].height = 20
    
    # Section header
    ws['B6'] = 'TO:'
    ws['B6'].font = header_font
    
    ws['B7'] = 'Alnoor Medical Services'
    ws['B7'].font = normal_font
    
    # Empty rows
    ws.row_dimensions[8].height = 5
    ws.row_dimensions[9].height = 5
    ws.row_dimensions[10].height = 5
    
    # Table Header Row 11
    ws['B11'] = 'ITEM DETAILS'
    ws['B11'].font = header_font
    ws.merge_cells('B11:F11')
    ws['B11'].alignment = center_align
    ws['B11'].fill = header_fill
    ws['B11'].border = thin_border
    
    # Column headers - Row 12
    headers = ['S.No', 'Item Code', 'Description', 'Qty (Ctns)', 'Qty/Ctn', 'Total Qty']
    header_cells = ['A12', 'C12', 'C12', 'D12', 'E12', 'F12']
    
    ws['A12'] = 'S.No'
    ws['C12'] = 'Item Code / Description'
    ws['D12'] = 'Qty (Ctns)'
    ws['E12'] = 'Qty/Ctn'
    ws['F12'] = 'Total Qty'
    
    for cell in ['A12', 'C12', 'D12', 'E12', 'F12']:
        ws[cell].font = header_font
        ws[cell].alignment = center_align
        ws[cell].fill = header_fill
        ws[cell].border = thin_border
    
    # Data Row 13 - Item Code (will be filled by program)
    ws['A13'] = '1'
    ws['A13'].alignment = center_align
    ws['A13'].border = thin_border
    
    ws['C13'] = ''  # Product reference will be filled here
    ws['C13'].alignment = left_align
    ws['C13'].border = thin_border
    
    # Data Row 14 - Item Description (will be filled by program)
    ws['C14'] = ''  # Product name will be filled here
    ws['C14'].alignment = left_align
    ws['C14'].border = thin_border
    
    ws['D14'] = 0  # Number of cartons (calculated)
    ws['D14'].alignment = center_align
    ws['D14'].border = thin_border
    
    ws['E14'] = 0  # Pieces per carton (user input)
    ws['E14'].alignment = center_align
    ws['E14'].border = thin_border
    
    ws['F14'] = 0  # Total pieces (from coupons)
    ws['F14'].alignment = center_align
    ws['F14'].border = thin_border
    
    # Empty rows
    ws.row_dimensions[15].height = 5
    ws.row_dimensions[16].height = 5
    ws.row_dimensions[17].height = 5
    ws.row_dimensions[18].height = 5
    
    # Total row - Row 19
    ws['E19'] = 'TOTAL:'
    ws['E19'].font = header_font
    ws['E19'].alignment = left_align
    
    ws['F20'] = 0  # Total (same as F14)
    ws['F20'].font = header_font
    ws['F20'].alignment = center_align
    ws['F20'].border = thin_border
    
    # Footer section
    ws.row_dimensions[21].height = 20
    ws.row_dimensions[22].height = 20
    
    ws['B22'] = 'Prepared By: _________________'
    ws['B22'].font = normal_font
    
    ws['E22'] = 'Received By: _________________'
    ws['E22'].font = normal_font
    
    ws['B24'] = 'Date: _________________'
    ws['B24'].font = normal_font
    
    ws['E24'] = 'Date: _________________'
    ws['E24'].font = normal_font
    
    # Save template
    template_path = Path(__file__).parent / "resources" / "templates" / "delivery_note_template.xlsx"
    template_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(template_path)
    print(f"Template created at: {template_path}")

if __name__ == "__main__":
    create_delivery_note_template()
